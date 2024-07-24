import openpyxl as xl

def excel_to_list(path, sheetName):
    file = xl.load_workbook(path, data_only=True)
    sheet = file[sheetName]
    data = []
    for row in sheet.iter_rows(values_only=True):
        # sheet를 가져와서 iter_rows선언할게 !
        # values_only를 안썼을 때는 주소까지 출력됨
        data.append(list(row))
    return data

경로 = "C:/Users/User/Desktop/seo/Python/파일만들기/월별구매리스트.xlsx"
new_경로 = "C:/Users/User/Desktop/seo/Python/파일만들기/새로운.xlsx"  # 새로 저장할 경로

# 원본 파일을 불러오기
file = xl.load_workbook(경로)

# 새 파일 생성
new_file = xl.Workbook()
new_sheet = new_file.active

# 첫 번째 시트의 이름 가져오기
li_sheetNames = file.sheetnames[:3]

# 첫 번째 시트에서 제목 가져오기 (2번 행)
시트0 = excel_to_list(경로, li_sheetNames[0])
제목 = 시트0[1]  # 2번 행은 인덱스 1

# 제목을 새 시트에 추가
new_sheet.append(제목)

# 각 시트의 데이터를 순회하며 새 시트에 추가
for 변수 in li_sheetNames:
    data1 = excel_to_list(경로, 변수)
    data1 = data1[2:]  # 0~1 행 제거 (제목 포함)

    for row in data1:  # 행 하나씩 꺼내서
        if any(row):  # 값이 하나라도 있는 경우
            new_sheet.append(row)

# 새 파일 저장
new_file.save(new_경로)
